import re
from datetime import datetime
import time
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Side, Border
from openpyxl.utils import get_column_letter

def extracted_messages(text, start_date, end_date):
    pattern = r'^(\d{2}/\d{2}/\d{2}), \d{2}:\d{2} - (Dinesh Padhye): (.+)$'
    messages = {}
    start_date_obj = datetime.strptime(start_date, "%d/%m/%y")
    end_date_obj = datetime.strptime(end_date, "%d/%m/%y")

    for line in text.split('\n'):
        match = re.match(pattern, line)
        if match:
            date, sender, message = match.groups()
            message = message.strip()
            message_is_upper = message.isupper()
            if start_date_obj <= datetime.strptime(date, "%d/%m/%y") <= end_date_obj:
                if message_is_upper and sender == 'Dinesh Padhye':
                    if date not in messages:
                        messages[date] = []
                    messages[date].append(message)

    return messages

with open("log.txt", "r", encoding='utf-8') as file:
    log_text = file.read()

start_date = input("Enter start date (DD/MM/YY): ")
end_date = input("Enter end date (DD/MM/YY): ")

messages = extracted_messages(log_text, start_date, end_date)

if messages:
    print("Messages sent by Dinesh Padhye in capital letters within the specified date range: \n")
    print(messages)
else:
    print("No messages found for the specified date range.")

#program for excel worksheet

wb = Workbook()
sheet = wb.active

month_input = int(input("Enter the month : "))

if month_input <= 9:
    heading_month = f"0{month_input}"
else:
    heading_month = month_input

krcl = "KONKAN RAILWAY CORPORATION LTD."
heading = f"Diary of MR.DINESH D. PADHYE (ESTM-I/RAJP) for the month of {heading_month}-2023, EMP. No. - 4203"

# first
sheet['A1'] = krcl
sheet['A1'].font = Font(name='Arial', size=20, bold=True, underline='single')
sheet.merge_cells('A1:H1')
sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')

# second
sheet['A4'] = heading
sheet['A4'].font = Font(name = 'Arial', size=15, bold=True, underline='single')
sheet.merge_cells('A4:H4')
sheet['A4'].alignment = Alignment(horizontal='center', vertical='center')

# chart mains
sheet['A6'] = 'DATE'
sheet['A6'].font = Font(name='Calibri', size=13, bold=True, italic=True, underline='single')
sheet.merge_cells('A6:A7')

sheet['B6'] = 'VISIT STATION'
sheet['B6'].font = Font(name='Calibri', size=13, bold=True, italic=True, underline='single')
sheet.merge_cells('B6:B7')

sheet['C6'] = 'ONWARD'
sheet['C6'].font = Font(name='Calibri', size=13, bold=True, italic=True, underline='single')
sheet.merge_cells('C6:D6')

sheet['C7'] = 'DEP.'
sheet['C7'].font = Font(name='Calibri', size=12, bold=True, italic=True, underline='single')

sheet['D7'] = 'ARR.'
sheet['D7'].font = Font(name='Calibri', size=12, bold=True, italic=True, underline='single')

sheet['E6'] = 'RETURN'
sheet['E6'].font = Font(name='Calibri', size=13, bold=True, italic=True, underline='single')
sheet.merge_cells('E6:F6')

sheet['E7'] = 'DEP.'
sheet['E7'].font = Font(name='Calibri', size=12, bold=True, italic=True, underline='single')

sheet['F7'] = 'ARR.'
sheet['F7'].font = Font(name='Calibri', size=12, bold=True, italic=True, underline='single')

sheet['G6'] = 'CONV.\nMODE'
sheet['G6'].font = Font(name='Calibri', size=13, bold=True, italic=True, underline='single')
sheet.merge_cells('G6:G7')

sheet['H6'] = 'NATURE OF WORK'
sheet['H6'].font = Font(name='Calibri', size=13, bold=True, italic=True, underline='single')
sheet.merge_cells('H6:H7')

# making centre alignment
for column in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
  for cell in sheet[column]:
    cell.alignment = Alignment(horizontal='center', vertical='center')

# setting values at page end
months = ["JAN.", "FEB.", "MAR.", "APR.", "MAY.", "JUN.", "JUL", "AUG", "SEP.", "OCT.", "NOV", "DEC"]

down_value = months[month_input]

sheet['A43'] = f"1. {down_value} 2023"
sheet['A43'].font = Font(name='Calibri', size=13, bold=True)

sheet['A44'] = f"RAJP"
sheet['A44'].font = Font(name='Calibri', size=13, bold=True)

sheet['H43'] = 'ESTM-I/RAJP'
sheet['H43'].font = Font(name='Calibri', size=13)

sheet['H44'] = 'DINESH D. PADHYE'
sheet['H44'].font = Font(name='Calibri', size=13, bold=True)

for row in range(43, 45):
    for column in range(1, 10):
        cell = sheet.cell(row=row, column=column)
        cell.alignment = Alignment(horizontal='center', vertical='center')

# adding main data

row_to_fill = 9

for date, message in messages.items():
    sheet.cell(row= row_to_fill, column=1).value = date
    sheet.cell(row= row_to_fill, column=8).value = "\n".join(message)
    row_to_fill = row_to_fill + 1

# adding dash for timing
cell_rang = 'C9:F36'

for row in sheet[cell_rang]:
  for cell in row:
    cell.value = '-'

# adding dash at conv mode
cells = 'G9:G36'

for row in sheet[cells]:
  for cell in row:
    cell.value = '-'

# Adding station

for row in range(1, 29):
  row = row + 8
  cell = sheet.cell(row=row, column=2)  # Column B is 2
  cell.value = 'RAJP'

# align data at center
cell_range = 'A9:G40'

for row in sheet[cell_range]:
  for cell in row:
    cell.alignment = Alignment(horizontal='center', vertical='center')

# align messages
for row in range(9, 41):
  cell = sheet.cell(row=row, column=8)
  cell.alignment = Alignment(vertical="center")

# adding border to heading
border_style = Side(style='thin')
border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

for row in range(1, 5):
  for col in range(1, 9):
    cell = sheet.cell(row=row, column=col)
    cell.border = border

# adding border to data
for row in range(6, 46):
  for col in range(1, 9):
    cell = sheet.cell(row=row, column=col)
    cell.border = border

# adjusting column widths
# for col in range(1, 9):
#     max_length = 0
#     for row in range(1, row_to_fill):
#         cell_value = str(sheet.cell(row=row, column=col).value)
#         max_length = max(max_length, len(cell_value))
#     adjusted_width = (max_length + 2) * 1.2
#     column_letter = get_column_letter(col)
#     sheet.column_dimensions[column_letter].width = adjusted_width


wb.save('ESTM Diary.xlsx')

print('\nSuccessfully created excel worksheet')